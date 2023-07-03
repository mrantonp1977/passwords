from tkinter import *
from PIL import ImageTk, Image
import sqlite3
import csv
from tkinter import ttk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook





root = Tk()
root.title("PASSWORD DATABASE")
icon = PhotoImage(file="logo.png")
root.iconphoto(True, icon)
root.geometry("600x500")
root.config(background="#0C847F")

conn = sqlite3.connect("antonis_passwords.db")
c = conn.cursor()

c.execute("""CREATE TABLE IF NOT EXISTS passwords (
file_app text,
username text,
password text)""")

app_combobox = ttk.Combobox(root, font=("arial black", 13), width=15)
app_combobox.grid(row=5, columnspan=2,  pady=10)


def update_app_list():
    global app_combobox
    c.execute("SELECT DISTINCT file_app FROM passwords")
    app = [row[0] for row in c.fetchall()]
    app_combobox['values'] = app



def update():
    conn = sqlite3.connect("antonis_passwords.db")
    c = conn.cursor()
    show_id = delete_box.get()
    c.execute("""UPDATE passwords SET
        file_app = :file_app,
        username = :username,
        password = :password
        
        WHERE oid = :oid""",
              {'file_app': file_app_editor.get(),
               'username': username_editor.get(),
               'password': password_editor.get(),
               "oid": show_id
               })

    confirmation = messagebox.askquestion("Επεξεργασία Αρχείων", "Do you want to save the changes  ???")
    if confirmation == 'yes':
        conn.commit()
        messagebox.showinfo(" Αποθήκευση Αρχείων ", "Επιτυχής αλλαγή και αποθήκευση Αρχείων. !!!")
    else:
        conn.rollback()
        messagebox.showinfo("Ακύρωση Εισαγωγής", "Η αλλαγή Αρχείων ακυρώθηκε. !!!")

    conn.commit()

    editor.destroy()

def edit():
    global editor
    editor = Tk()
    editor.title("ΕΠΕΞΕΡΓΑΣΙΑ ΣΤΟΙΧΕΙΩΝ")
    editor.geometry("600x300")
    editor.config(background="#F3F3C3")

    conn = sqlite3.connect("antonis_passwords.db")
    c = conn.cursor()

    show_id = delete_box.get()

    c.execute("SELECT * FROM passwords WHERE oid = " + show_id)
    shows = c.fetchall()

    global file_app_editor
    global username_editor
    global password_editor


    file_app_editor = Entry(editor, width=20, font=("helvetica", 17))
    file_app_editor.grid(row=0, column=1, padx=50, pady=(10, 0))
    username_editor = Entry(editor, width=20, font=("helvetica", 17))
    username_editor.grid(row=1, column=1, padx=50, pady=(10, 0))
    password_editor = Entry(editor, width=20, font=("helvetica", 17))
    password_editor.grid(row=2, column=1, padx=50, pady=(10, 0))


    file_app_label = Label(editor, text="ΕΦΑΡΜΟΓΗ :", font=("arial black", 16), background="#F3F3C3", foreground="#7E0B9E")
    file_app_label.grid(row=0, column=0, pady=(10, 0), padx=15, sticky=W)
    username_label = Label(editor, text="USERNAME :", font=("arial black", 16), background="#F3F3C3", foreground="#7E0B9E")
    username_label.grid(row=1, column=0, sticky=W, pady=10, padx=15)
    password_label = Label(editor, text="PASSWORD :", font=("arial black", 16), background="#F3F3C3", foreground="#7E0B9E")
    password_label.grid(row=2, column=0, sticky=W, pady=10, padx=15)

    for show in shows:
        file_app_editor.insert(0, show[0])
        username_editor.insert(0, show[1])
        password_editor.insert(0, show[2])


    edit_btn = Button(editor, text="Επεξεργασία και Αποθήκευση \n Στοιχείων", command=update, font=("arial black", 12), background="#06D784", activebackground="#06D784", activeforeground="#06D784")
    edit_btn.grid(row=3, column=0, columnspan=2, pady=10, padx=2, ipadx=10)





def delete():
    conn = sqlite3.connect("antonis_passwords.db")
    c = conn.cursor()

    c.execute("DELETE from passwords WHERE oid = " + delete_box.get())



    conn.commit()
    conn.close()

def submit():
    conn = sqlite3.connect("antonis_passwords.db")
    c = conn.cursor()




    c.execute("INSERT INTO passwords VALUES (:file_app, :username, :password)",
              {
                  "file_app": file_app.get(),
                  "username": username.get(),
                  "password": password.get(),

              })

    confirmation = messagebox.askquestion("Εισαγωγή Αρχείων", "Do You Want To Insert this Files ???")
    if confirmation == 'yes':
        conn.commit()
        messagebox.showinfo(" Εισαγωγή Αρχείων ", "Επιτυχής Εισαγωγή Αρχείων. !!!")
    else:
        conn.rollback()
        messagebox.showinfo("Ακύρωση Εισαγωγής", "Η Εισαγωγή Αρχείων ακυρώθηκε. !!!")
    conn.close()

    file_app.delete(0, END)
    username.delete(0, END)
    password.delete(0, END)

def save_to_excel():
    conn = sqlite3.connect("antonis_passwords.db")
    c = conn.cursor()

    c.execute("SELECT * FROM passwords")
    data = c.fetchall()

    try:
        workbook = load_workbook("total.apps.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        headers = ["ΟΜΟΜΑ ΕΦΑΡΜΟΓΗΣ", "USERNAME", "PASSWORD"]
        sheet.append(headers)

    # Clear existing data in the worksheet
    sheet.delete_rows(2, sheet.max_row)

    for row in data:
        sheet.append(row)

    workbook.save("total.apps.xlsx")

    conn.close()

    messagebox.showinfo("Αποθήκευση Δεδομένων", "Τα δεδομένα αποθηκεύτηκαν σε αρχείο Excel !!!")

def delete_all_data():
    confirmation = messagebox.askyesno("Confirmation", "Είστε σίγουρος ότι θέλετε να διαγράψετε όλα τα Αρχεία ;")
    if confirmation:
        conn = sqlite3.connect("antonis_passwords.db")
        c = conn.cursor()
        c.execute("DELETE FROM passwords")
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", "Ολα τα Αρχεία διαγράφηκαν με επιτυχία !!")


def show():
    show = Tk()
    show.title("PASSWORD DATABASE")
    show.geometry("600x700")
    show.config(background="#F3F3C3")

    conn = sqlite3.connect("antonis_passwords.db")
    c = conn.cursor()

    c.execute("SELECT *, oid FROM passwords")
    elements = c.fetchall()

    for i, element in enumerate(elements):
        text = '    '.join(str(item) for item in element)
        show_label = Label(show, text=text, font=("arial black", 11), background="#F3F3C3", foreground="#7E0B9E")
        show_label.grid(row=i, column=0, pady=5, padx=20, sticky='w')

    save_button = Button(show, text="Αποθήκευση σε 'Excel' ", command=save_to_excel, font=("arial black", 11),
                         background="#06D784")
    save_button.grid(row=i + 1, column=0, pady=10, ipadx=26, sticky=W, padx=20)
    delete_button = Button(show, text="Διαγραφή όλων των Αρχείων", command=delete_all_data, font=("arial black", 11),
                           background="#EA6969")
    delete_button.grid(row=i + 1, column=0, pady=10, padx=320, ipadx=8)






def select_app():
    selected_app = app_combobox.get()
    if selected_app:
        c.execute("SELECT file_app, username, password FROM passwords WHERE file_app=? GROUP BY file_app", (selected_app,))
        app_info = c.fetchall()
        if app_info:
            app_window = Toplevel(root)
            app_window.title("Worker Information")
            app_window.geometry("550x300")
            app_window.config(background="#0C847F")

            info_labels = [" ΌΝΟΜΑ ΕΦΑΡΜΟΓΗΣ : ", " USERNAME : ", " PASSWORD : "]

            for row_index, worker_data in enumerate(app_info):
                for i, label in enumerate(info_labels):
                    label_text = Label(app_window, text=label, font=("arial black", 12), width=20, relief="ridge", bd=3,  background="#F1B52D")
                    label_text.grid(row=row_index * len(info_labels) + i, column=0, padx=20, sticky=W, pady=10)

                    if i == 1:
                        # Display the sum of prices
                        value_label = Label(app_window, text=worker_data[i], font=("arial black", 13), foreground="yellow",  background="#0C847F")
                    else:
                        # Display the worker name
                        value_label = Label(app_window, text=worker_data[i], font=("arial black", 13), foreground="yellow",  background="#0C847F")

                    value_label.grid(row=row_index * len(info_labels) + i, column=1, padx=20, pady=10)

        else:
            messagebox.showerror("Error", "No information available for the selected worker.")
    else:
        messagebox.showwarning("Warning", "Please select a worker.")

    update_app_list()



update_app_list()


file_app = Entry(root, width=20, font=("helvetica", 14))
file_app.grid(row=1, column=1, padx=50, pady=(10, 0))
username = Entry(root, width=20, font=("helvetica", 14))
username.grid(row=2, column=1, padx=50, pady=(10, 0))
password = Entry(root, width=20, font=("helvetica", 14))
password.grid(row=3, column=1, padx=50, pady=(10, 0))
delete_box = Entry(root, width=5, font=("helvetica", 14))
delete_box.grid(row=4, column=1, padx=50, pady=(10, 0))

file_app_label = Label(root, text="ΕΦΑΡΜΟΓΗ :", font=("arial black", 12), background="#0C847F",  foreground="#E0E931")
file_app_label.grid(row=1, column=0, pady=(10, 0), padx=15, sticky=W)
username_label = Label(root, text="USERNAME :", font=("arial black", 12), background="#0C847F", foreground="#E0E931")
username_label.grid(row=2, column=0, sticky=W, pady=10, padx=15)
password_label = Label(root, text="PASSWORD :", font=("arial black", 12), background="#0C847F", foreground="#E0E931")
password_label.grid(row=3, column=0, sticky=W, pady=10, padx=15)
delete_box_label = Label(root, text="ID NUMBER :", font=("arial black", 12), background="#0C847F", foreground="#E0E931")
delete_box_label.grid(row=4, column=0, sticky=W, pady=10, padx=15)

title_label = Label(root, text="          MY PASSWORDS          ", font=("IMPACT", 20), background="black", relief="ridge", borderwidth=10, foreground="yellow")
title_label.grid(row=0, column=0, columnspan=2, pady=20, padx=25)


submit_btn = Button(root, text="Εισαγωγή Στοιχείων", command=submit, bd=4, font=("arial black", 11),  background="#06D784", activebackground="#06D784", activeforeground="#06D784")
submit_btn.grid(row=6, column=0, pady=20, padx=10, ipadx=30)

show_btn = Button(root, text="Εμφάνιση Στοιχείων", command=show, bd=4, font=("arial black", 11), background="#6ECDF5", activebackground="#6ECDF5", activeforeground="#6ECDF5")
show_btn.grid(row=6, column=1,  pady=10, padx=10, ipadx=30)

delete_btn = Button(root, text="Διαγραφή Στοιχείων", command=delete, bd=4, font=("arial black", 11), background="#F67EA3", activebackground="#F67EA3", activeforeground="#F67EA3")
delete_btn.grid(row=7, column=0,  pady=10, padx=10, ipadx=30)

edit_btn = Button(root, text="Επεξεργασία Στοιχείων", command=edit, bd=4, font=("arial black", 11), background="#A481C9", activebackground="#A481C9", activeforeground="#A481C9")
edit_btn.grid(row=7, column=1,  pady=10, padx=10, ipadx=20)

select_btn = Button(root, text="Επιλογή Εφαρμογής :", command=select_app, font=("arial black", 9), background="blue", foreground="yellow")
select_btn.grid(row=5, column=0, sticky=W, padx=10, pady=10)






conn.commit()

root.mainloop()